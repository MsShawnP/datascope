"""Tests for datascope.reports.annotated_excel -- annotated Excel output."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from datascope.findings.composer import compose_finding
from datascope.findings.pipeline import process_findings
from datascope.findings.severity import classify_severity
from datascope.models import Finding, FindingType
from datascope.reports.annotated_excel import write_annotated_excel

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_DEFAULT_METADATA = {
    "filename": "test_data.csv",
    "row_count": 5,
    "column_count": 3,
}

_DEFAULT_HEADERS = ["id", "revenue", "status"]

_DEFAULT_DATA = [
    [1, 100.0, "Active"],
    [2, 200.0, "Active"],
    [3, "N/A", "Inactive"],
    [4, 400.0, "Active"],
    [5, 500.0, "Active"],
]


def _make_finding(
    finding_type: FindingType,
    evidence: dict,
    field_name: str = "test_col",
) -> Finding:
    f = Finding(field_name=field_name, finding_type=finding_type, evidence=evidence)
    f.severity = classify_severity(f)
    compose_finding(f)
    return f


def _critical_finding(field_name: str = "revenue") -> Finding:
    return _make_finding(
        FindingType.TYPE_INCONSISTENCY,
        {
            "majority_type": "numeric",
            "majority_count": 4,
            "minority_types": [
                {"type_name": "str", "count": 1, "examples": ["N/A"]},
            ],
            "total_non_null": 5,
            "majority_pct": 80.0,
        },
        field_name=field_name,
    )


def _warning_finding(field_name: str = "zip_code") -> Finding:
    return _make_finding(
        FindingType.LEADING_ZEROS,
        {
            "leading_zero_count": 10,
            "no_leading_zero_count": 40,
            "examples_with_zeros": ["00123", "00456", "00789"],
            "examples_without_zeros": ["123", "456"],
            "total_checked": 50,
        },
        field_name=field_name,
    )


def _info_finding(field_name: str = "status") -> Finding:
    return _make_finding(
        FindingType.NEAR_CONSTANT,
        {
            "unique_count": 2,
            "total_count": 5,
            "uniqueness_ratio": 0.4,
            "top_values": [
                {"value": "Active", "count": 4},
                {"value": "Inactive", "count": 1},
            ],
        },
        field_name=field_name,
    )


# ---------------------------------------------------------------------------
# Happy path -- mixed findings
# ---------------------------------------------------------------------------

class TestHappyPathMixedFindings:
    @pytest.fixture()
    def wb_path(self, tmp_path: Path) -> Path:
        findings = process_findings([
            _critical_finding("revenue"),
            _info_finding("status"),
        ])
        out = tmp_path / "annotated.xlsx"
        write_annotated_excel(findings, _DEFAULT_METADATA, _DEFAULT_DATA, _DEFAULT_HEADERS, out)
        return out

    def test_file_created(self, wb_path: Path):
        assert wb_path.exists()

    def test_file_non_zero_size(self, wb_path: Path):
        assert wb_path.stat().st_size > 0

    def test_has_data_sheet(self, wb_path: Path):
        wb = load_workbook(wb_path)
        assert "Data" in wb.sheetnames

    def test_has_findings_sheet(self, wb_path: Path):
        wb = load_workbook(wb_path)
        assert "Findings" in wb.sheetnames

    def test_data_sheet_headers(self, wb_path: Path):
        wb = load_workbook(wb_path)
        ws = wb["Data"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, 4)]
        assert headers == ["id", "revenue", "status"]

    def test_data_sheet_row_count(self, wb_path: Path):
        wb = load_workbook(wb_path)
        ws = wb["Data"]
        assert ws.max_row == 6  # 1 header + 5 data rows

    def test_data_values_preserved(self, wb_path: Path):
        wb = load_workbook(wb_path)
        ws = wb["Data"]
        assert ws.cell(row=2, column=1).value == 1
        assert ws.cell(row=2, column=2).value == 100.0
        assert ws.cell(row=4, column=3).value == "Inactive"

    def test_findings_sheet_headers(self, wb_path: Path):
        wb = load_workbook(wb_path)
        ws = wb["Findings"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, 7)]
        assert headers == ["Field", "Issue Type", "Severity", "Assumption", "Reality", "Fix"]

    def test_findings_sheet_row_count(self, wb_path: Path):
        wb = load_workbook(wb_path)
        ws = wb["Findings"]
        assert ws.max_row == 3  # 1 header + 2 findings

    def test_finding_field_names(self, wb_path: Path):
        wb = load_workbook(wb_path)
        ws = wb["Findings"]
        fields = [ws.cell(row=r, column=1).value for r in range(2, 4)]
        assert "revenue" in fields
        assert "status" in fields


# ---------------------------------------------------------------------------
# Header styling
# ---------------------------------------------------------------------------

class TestHeaderStyling:
    @pytest.fixture()
    def wb(self, tmp_path: Path):
        findings = process_findings([_critical_finding("revenue")])
        out = tmp_path / "styled.xlsx"
        write_annotated_excel(findings, _DEFAULT_METADATA, _DEFAULT_DATA, _DEFAULT_HEADERS, out)
        return load_workbook(out)

    def test_data_header_fill(self, wb):
        ws = wb["Data"]
        fill = ws.cell(row=1, column=1).fill
        assert fill.start_color.rgb is not None

    def test_data_header_font_color(self, wb):
        ws = wb["Data"]
        font = ws.cell(row=1, column=1).font
        assert font.bold is True
        assert font.color.rgb == "00FFFFFF"

    def test_findings_header_fill(self, wb):
        ws = wb["Findings"]
        fill = ws.cell(row=1, column=1).fill
        assert fill.start_color.rgb is not None

    def test_findings_header_font(self, wb):
        ws = wb["Findings"]
        font = ws.cell(row=1, column=1).font
        assert font.bold is True


# ---------------------------------------------------------------------------
# Severity cell coloring
# ---------------------------------------------------------------------------

class TestSeverityColoring:
    @pytest.fixture()
    def wb(self, tmp_path: Path):
        findings = process_findings([
            _critical_finding("revenue"),
            _info_finding("status"),
        ])
        out = tmp_path / "colors.xlsx"
        write_annotated_excel(findings, _DEFAULT_METADATA, _DEFAULT_DATA, _DEFAULT_HEADERS, out)
        return load_workbook(out)

    def test_flagged_column_has_fill(self, wb):
        ws = wb["Data"]
        cell = ws.cell(row=2, column=2)  # revenue column, first data row
        assert cell.fill.start_color.rgb != "00000000"

    def test_unflagged_column_no_fill(self, wb):
        ws = wb["Data"]
        cell = ws.cell(row=2, column=1)  # id column, not flagged
        assert cell.fill.start_color.rgb == "00000000"

    def test_critical_uses_critical_tint(self, wb):
        ws = wb["Data"]
        cell = ws.cell(row=2, column=2)  # revenue = critical
        assert cell.fill.start_color.rgb == "00fce8e7"

    def test_info_uses_info_tint(self, wb):
        ws = wb["Data"]
        cell = ws.cell(row=2, column=3)  # status = info
        assert cell.fill.start_color.rgb == "00e8eaf4"

    def test_worst_severity_wins(self, tmp_path: Path):
        findings = process_findings([
            _critical_finding("revenue"),
            _info_finding("revenue"),
        ])
        out = tmp_path / "worst.xlsx"
        write_annotated_excel(findings, _DEFAULT_METADATA, _DEFAULT_DATA, _DEFAULT_HEADERS, out)
        wb = load_workbook(out)
        ws = wb["Data"]
        cell = ws.cell(row=2, column=2)
        assert cell.fill.start_color.rgb == "00fce8e7"


# ---------------------------------------------------------------------------
# Finding content on Findings sheet
# ---------------------------------------------------------------------------

class TestFindingsContent:
    @pytest.fixture()
    def ws(self, tmp_path: Path):
        findings = process_findings([_critical_finding("revenue")])
        out = tmp_path / "content.xlsx"
        write_annotated_excel(findings, _DEFAULT_METADATA, _DEFAULT_DATA, _DEFAULT_HEADERS, out)
        wb = load_workbook(out)
        return wb["Findings"]

    def test_field_name(self, ws):
        assert ws.cell(row=2, column=1).value == "revenue"

    def test_issue_type_label(self, ws):
        assert ws.cell(row=2, column=2).value == "Type Inconsistency"

    def test_severity_label(self, ws):
        assert ws.cell(row=2, column=3).value == "Critical"

    def test_assumption_populated(self, ws):
        val = ws.cell(row=2, column=4).value
        assert val is not None and len(val) > 0

    def test_reality_populated(self, ws):
        val = ws.cell(row=2, column=5).value
        assert val is not None and len(val) > 0

    def test_fix_populated(self, ws):
        val = ws.cell(row=2, column=6).value
        assert val is not None and len(val) > 0

    def test_findings_row_fill(self, ws):
        fill = ws.cell(row=2, column=1).fill
        assert fill.start_color.rgb == "00fce8e7"


# ---------------------------------------------------------------------------
# Zero findings
# ---------------------------------------------------------------------------

class TestZeroFindings:
    def test_file_created(self, tmp_path: Path):
        out = tmp_path / "empty.xlsx"
        write_annotated_excel([], _DEFAULT_METADATA, _DEFAULT_DATA, _DEFAULT_HEADERS, out)
        assert out.exists()

    def test_data_still_written(self, tmp_path: Path):
        out = tmp_path / "empty.xlsx"
        write_annotated_excel([], _DEFAULT_METADATA, _DEFAULT_DATA, _DEFAULT_HEADERS, out)
        wb = load_workbook(out)
        ws = wb["Data"]
        assert ws.max_row == 6

    def test_findings_sheet_header_only(self, tmp_path: Path):
        out = tmp_path / "empty.xlsx"
        write_annotated_excel([], _DEFAULT_METADATA, _DEFAULT_DATA, _DEFAULT_HEADERS, out)
        wb = load_workbook(out)
        ws = wb["Findings"]
        assert ws.max_row == 1


# ---------------------------------------------------------------------------
# Empty data (headers only, no rows)
# ---------------------------------------------------------------------------

class TestEmptyData:
    def test_headers_only(self, tmp_path: Path):
        out = tmp_path / "headers_only.xlsx"
        write_annotated_excel([], _DEFAULT_METADATA, [], _DEFAULT_HEADERS, out)
        wb = load_workbook(out)
        ws = wb["Data"]
        assert ws.max_row == 1
        assert ws.cell(row=1, column=1).value == "id"


# ---------------------------------------------------------------------------
# Auto-create parent directory
# ---------------------------------------------------------------------------

class TestAutoCreateDirectory:
    def test_creates_nested_directory(self, tmp_path: Path):
        out = tmp_path / "sub" / "deep" / "annotated.xlsx"
        write_annotated_excel([], _DEFAULT_METADATA, _DEFAULT_DATA, _DEFAULT_HEADERS, out)
        assert out.exists()


# ---------------------------------------------------------------------------
# Column widths
# ---------------------------------------------------------------------------

class TestColumnWidths:
    def test_data_columns_have_width(self, tmp_path: Path):
        out = tmp_path / "widths.xlsx"
        write_annotated_excel([], _DEFAULT_METADATA, _DEFAULT_DATA, _DEFAULT_HEADERS, out)
        wb = load_workbook(out)
        ws = wb["Data"]
        assert ws.column_dimensions["A"].width == 18

    def test_findings_columns_have_varied_widths(self, tmp_path: Path):
        findings = process_findings([_critical_finding()])
        out = tmp_path / "widths.xlsx"
        write_annotated_excel(findings, _DEFAULT_METADATA, _DEFAULT_DATA, _DEFAULT_HEADERS, out)
        wb = load_workbook(out)
        ws = wb["Findings"]
        assert ws.column_dimensions["A"].width == 18
        assert ws.column_dimensions["B"].width == 22
        assert ws.column_dimensions["C"].width == 10
