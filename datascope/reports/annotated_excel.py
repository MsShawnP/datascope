"""Annotated Excel report generator.

Copies the source data into a new workbook, highlights problem cells with
conditional formatting (red/amber/blue by severity), and adds a "Findings"
summary sheet.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from datascope.models import Finding, Severity
from datascope.reports._palette import (
    CHICAGO_20_HEX,
    CRITICAL_TINT_HEX,
    FINDING_TYPE_LABELS,
    INFO_TINT_HEX,
    SEVERITY_LABELS,
    WARNING_TINT_HEX,
)

_SEVERITY_FILLS = {
    Severity.CRITICAL: PatternFill(start_color=CRITICAL_TINT_HEX[1:], end_color=CRITICAL_TINT_HEX[1:], fill_type="solid"),
    Severity.WARNING: PatternFill(start_color=WARNING_TINT_HEX[1:], end_color=WARNING_TINT_HEX[1:], fill_type="solid"),
    Severity.INFO: PatternFill(start_color=INFO_TINT_HEX[1:], end_color=INFO_TINT_HEX[1:], fill_type="solid"),
}

_HEADER_FILL = PatternFill(start_color=CHICAGO_20_HEX[1:], end_color=CHICAGO_20_HEX[1:], fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)


def write_annotated_excel(
    findings: list[Finding],
    source_metadata: dict[str, Any],
    source_data: list[list[Any]],
    headers: list[str],
    output_path: Path,
) -> None:
    """Write an annotated Excel workbook.

    Parameters
    ----------
    findings:
        Processed findings with severity and narrative text.
    source_metadata:
        Source metadata dict.
    source_data:
        Row-major list of lists with the original cell values.
    headers:
        Column headers from the source data.
    output_path:
        Path for the output .xlsx file.
    """
    wb = Workbook()

    # --- Data sheet with highlights ------------------------------------
    ws_data = wb.active
    ws_data.title = "Data"

    field_to_severity: dict[str, Severity] = {}
    for f in findings:
        sev = f.severity or Severity.INFO
        existing = field_to_severity.get(f.field_name)
        if existing is None or sev > existing:
            field_to_severity[f.field_name] = sev

    for col_idx, header in enumerate(headers, start=1):
        cell = ws_data.cell(row=1, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT

    for row_idx, row_data in enumerate(source_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_data.cell(row=row_idx, column=col_idx, value=value)
            if col_idx <= len(headers):
                header = headers[col_idx - 1]
                if header in field_to_severity:
                    cell.fill = _SEVERITY_FILLS[field_to_severity[header]]

    for col_idx in range(1, len(headers) + 1):
        ws_data.column_dimensions[get_column_letter(col_idx)].width = 18

    # --- Findings sheet ------------------------------------------------
    ws_findings = wb.create_sheet("Findings")

    finding_headers = ["Field", "Issue Type", "Severity", "Assumption", "Reality", "Fix"]
    for col_idx, header in enumerate(finding_headers, start=1):
        cell = ws_findings.cell(row=1, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT

    for row_idx, f in enumerate(findings, start=2):
        sev = f.severity or Severity.INFO
        fill = _SEVERITY_FILLS[sev]
        type_label = FINDING_TYPE_LABELS.get(f.finding_type, f.finding_type.value)
        sev_label = SEVERITY_LABELS.get(sev, "Info")

        values = [
            f.field_name,
            type_label,
            sev_label,
            f.assumption or "",
            f.reality or "",
            f.fix_recommendation or "",
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws_findings.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill

    col_widths = [18, 22, 10, 40, 50, 40]
    for col_idx, width in enumerate(col_widths, start=1):
        ws_findings.column_dimensions[get_column_letter(col_idx)].width = width

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
