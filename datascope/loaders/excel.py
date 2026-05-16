"""Excel loader -- cell-level type-preserving reader using openpyxl.

Extracted from the proven ``load_strict()`` pattern in ``_legacy.py``.
Returns a :class:`~datascope.models.LoaderResult` so that downstream
analysers never need to know which loader ran.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd

from datascope.models import LoaderResult


def load_excel(path: Path, sheet: str | int = 0) -> LoaderResult:
    """Read an ``.xlsx`` workbook cell-by-cell, preserving native Python types.

    Uses openpyxl with ``data_only=True`` (formula cells return their
    last-cached value, not re-evaluated) and ``read_only=True``
    (memory-efficient streaming).

    Parameters
    ----------
    path:
        Path to the ``.xlsx`` file.
    sheet:
        Sheet name (``str``) or 0-based index (``int``).

    Returns
    -------
    LoaderResult
        DataFrame with ``dtype=object``, per-column ``cell_types``, and
        source metadata.
    """
    from openpyxl import load_workbook

    path = Path(path)
    wb = load_workbook(path, data_only=True, read_only=True)

    try:
        if isinstance(sheet, int):
            if sheet < 0 or sheet >= len(wb.worksheets):
                available = [ws.title for ws in wb.worksheets]
                wb.close()
                raise ValueError(
                    f"Sheet index {sheet} is out of range. "
                    f"Available sheets: {', '.join(available)}"
                )
            ws = wb.worksheets[sheet]
        else:
            if sheet not in wb.sheetnames:
                available = wb.sheetnames
                wb.close()
                raise ValueError(
                    f"Sheet '{sheet}' not found. "
                    f"Available sheets: {', '.join(available)}"
                )
            ws = wb[sheet]

        sheet_title = ws.title
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    # Empty workbook / empty sheet
    if not rows:
        return LoaderResult(
            dataframe=pd.DataFrame(),
            cell_types={},
            source_metadata={
                "filename": path.name,
                "sheet": sheet_title,
                "row_count": 0,
                "column_count": 0,
            },
        )

    # --- headers --------------------------------------------------------
    headers = [
        str(h) if h is not None else f"col_{i}"
        for i, h in enumerate(rows[0])
    ]

    # --- data rows (pad jagged rows to header width) ---------------------
    n_cols = len(headers)
    data_rows = rows[1:]
    data = [list(r) + [None] * max(0, n_cols - len(r)) for r in data_rows]

    df = pd.DataFrame(data, columns=headers, dtype=object)

    # --- cell_types -----------------------------------------------------
    cell_types: dict[str, list[type]] = {}
    for col_idx, col_name in enumerate(headers):
        cell_types[col_name] = [
            type(row[col_idx]) for row in data
        ]

    # --- source_metadata ------------------------------------------------
    source_metadata = {
        "filename": path.name,
        "sheet": sheet_title,
        "row_count": len(data_rows),
        "column_count": len(headers),
    }

    return LoaderResult(
        dataframe=df,
        cell_types=cell_types,
        source_metadata=source_metadata,
    )
